"""Comp import — parse MLS exports (CSV/Excel/PDF) into structured comp records.

Supports three input formats:
  - CSV (from Matrix, Flexmls, etc.)
  - Excel (.xlsx, .xls, .xlsm)
  - PDF (Comparable Sales Reports from data providers like PropStream)

Returns a list of comp dicts with normalized keys.
"""
import re, csv, io
from typing import Optional, List, Dict, Any


COLUMN_PATTERNS = {
    "address":   ["property address", "full address", "site address", "street address", "address"],
    "city":      ["city", "municipality"],
    "state":     ["state"],
    "zip":       ["zip code", "zipcode", "postal code", "postal", "zip"],
    "beds":      ["bedrooms", "bed rooms", "beds", "br", "# beds"],
    "baths":     ["total baths", "bathrooms", "baths", "ba", "# baths", "full baths"],
    "sqft":      ["total living area", "living area sqft", "living sqft", "living area",
                  "sq ft", "sqft", "square feet", "approx sqft"],
    "year":      ["year built", "yr built", "year"],
    "sold_price":["close price", "closed price", "sold price", "sale price", "sold for", "price"],
    "sold_date": ["close date", "closed date", "sold date", "sale date", "closing date"],
    "distance":  ["approximate distance", "distance", "miles", "mi"],
    "notes":     ["public remarks", "remarks", "notes", "comments"],
}


def _normalize(s):
    if s is None: return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _find_column(header, patterns):
    norm = [_normalize(h) for h in header]
    for p in patterns:
        pn = _normalize(p)
        for i, h in enumerate(norm):
            if h == pn: return i
    for p in patterns:
        pn = _normalize(p)
        for i, h in enumerate(norm):
            if pn and (pn in h or (len(h) >= 3 and h in pn)):
                return i
    return None


def _parse_money(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try: return float(s) if s else None
    except ValueError: return None


def _parse_int(v):
    n = _parse_money(v)
    return int(n) if n is not None else None


def _parse_first_int(v):
    """Extract the FIRST integer from a string. Handles 'X,XXX Δ -YYY' format."""
    if v is None or v == "": return None
    if isinstance(v, int): return v
    if isinstance(v, float): return int(v)
    m = re.search(r"[\d,]+", str(v))
    if not m: return None
    try: return int(m.group(0).replace(",", ""))
    except ValueError: return None


def _read_csv_bytes(data: bytes) -> List[List]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_excel_bytes(data: bytes) -> List[List]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ============================================================================
# PDF PARSER — coordinate-based, handles column-tabular CMA reports
# ============================================================================
def _cluster_x_positions(xs, tolerance=60):
    if not xs:
        return []
    sorted_xs = sorted(set(xs))
    clusters = [[sorted_xs[0]]]
    for x in sorted_xs[1:]:
        if x - clusters[-1][-1] <= tolerance:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c) / len(c) for c in clusters]


def _assign_to_column(x, column_anchors, max_extent=120):
    """Return index of column whose anchor is just to the left of x."""
    if not column_anchors:
        return None
    if x < column_anchors[0] - 10:
        return None
    for i in range(len(column_anchors) - 1):
        if column_anchors[i] - 10 <= x < column_anchors[i+1] - 10:
            return i
    last = column_anchors[-1]
    if x >= last - 10 and x < last + max_extent:
        return len(column_anchors) - 1
    return None


def _parse_pdf_comps(data: bytes) -> List[Dict[str, Any]]:
    """Parse a Comparable Sales Report style PDF (from PropStream and similar)."""
    import pdfplumber
    all_comps = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "Comparable List" not in text and "Comparable Sales Report" not in text:
                continue
            words = page.extract_words(use_text_flow=True)
            if not words:
                continue

            from collections import defaultdict
            rows = defaultdict(list)
            for w in words:
                key = round(w["top"] / 5) * 5
                rows[key].append(w)

            candidate_xs = []
            for y in sorted(rows):
                line_words = sorted(rows[y], key=lambda w: w["x0"])
                if not line_words: continue
                first_text = line_words[0]["text"]
                if first_text in ("Bed", "Bath", "Stories", "AVM"):
                    for w in line_words[1:]:
                        candidate_xs.append(w["x0"])

            if not candidate_xs:
                continue

            column_centers = _cluster_x_positions(candidate_xs, tolerance=60)
            if len(column_centers) < 2:
                continue

            fields = defaultdict(lambda: defaultdict(list))
            current_field = None
            for y in sorted(rows):
                line_words = sorted(rows[y], key=lambda w: w["x0"])
                if not line_words: continue
                first_word = line_words[0]
                first_text = first_word["text"]
                first_x = first_word["x0"]

                if first_x < column_centers[0] - 20:
                    label_words = []
                    value_words = []
                    for w in line_words:
                        if w["x0"] < column_centers[0] - 20:
                            label_words.append(w["text"])
                        else:
                            value_words.append(w)
                    label = " ".join(label_words).strip()
                    current_field = label
                    for w in value_words:
                        col = _assign_to_column(w["x0"], column_centers)
                        if col is not None:
                            fields[current_field][col].append(w["text"])
                else:
                    if current_field:
                        for w in line_words:
                            col = _assign_to_column(w["x0"], column_centers)
                            if col is not None:
                                fields[current_field][col].append(w["text"])

            num_comps = len(column_centers) - 1
            for comp_idx in range(num_comps):
                col = comp_idx + 1
                def get(field_name):
                    tokens = fields.get(field_name, {}).get(col, [])
                    return " ".join(tokens).strip() if tokens else None

                last_sold = get("Last Sold")
                price = None
                date = None
                if last_sold:
                    m = re.search(r"\$[\d,]+", last_sold)
                    if m: price = _parse_money(m.group())
                    m = re.search(r"\d{1,2}/\d{1,2}/\d{4}", last_sold)
                    if m: date = m.group()

                sqft = _parse_first_int(get("Square Feet"))
                beds = _parse_first_int(get("Bed"))
                baths = get("Bath")
                year = _parse_int(get("Year Built"))
                listing_status = get("Listing Status")

                if price is None:
                    continue

                comp = {
                    "address": None, "city": None, "state": None, "zip": None,
                    "beds": beds, "baths": baths, "sqft": sqft, "year": year,
                    "sold_price": price, "sold_date": date,
                    "distance": None, "notes": listing_status or "",
                    "dollar_per_sqft": (price / sqft) if (price and sqft) else None,
                }
                all_comps.append((col, comp))

            # Extract addresses from rows above "Bed"
            sorted_ys = sorted(rows)
            bed_y = None
            for y in sorted_ys:
                texts = [w["text"] for w in rows[y]]
                if "Bed" in texts:
                    bed_y = y; break

            if bed_y is not None:
                addr_rows = []
                for y in sorted_ys:
                    if y >= bed_y: continue
                    texts = [w["text"] for w in rows[y]]
                    if any(t in ("Subject", "Property", "Comparable", "List") for t in texts):
                        continue
                    addr_rows.append(y)
                addr_rows = addr_rows[-4:] if len(addr_rows) > 4 else addr_rows

                addr_by_col = defaultdict(list)
                for y in addr_rows:
                    for w in sorted(rows[y], key=lambda w: w["x0"]):
                        col = _assign_to_column(w["x0"], column_centers)
                        if col is not None and col >= 1:
                            addr_by_col[col].append(w["text"])

                for col, addr_tokens in addr_by_col.items():
                    full = " ".join(addr_tokens)
                    zip_matches = re.findall(r"\b(\d{5})\b", full)
                    zip_code = zip_matches[-1] if zip_matches else None
                    state_code = None
                    state_pattern = (r"\b(FL|GA|NC|SC|TX|CA|NY|NJ|PA|VA|MD|MA|TN|OH|"
                                     r"MI|IL|IN|KY|AL|AR|LA|MS|MO|OK|KS|CO|UT|AZ|NV|"
                                     r"OR|WA|WI|MN|IA|NE|SD|ND|MT|WY|ID|NM|HI|AK|ME|"
                                     r"VT|NH|RI|CT|DE|WV)\b\s+\d{5}\b")
                    sm = re.search(state_pattern, full, re.I)
                    if sm: state_code = sm.group(1).upper()
                    for c2, comp in all_comps:
                        if c2 == col and comp.get("address") is None:
                            comp["address"] = full
                            comp["zip"] = zip_code
                            comp["state"] = state_code
                            break

    return [c for _, c in all_comps]


def parse_comp_file(file_obj_or_bytes, filename: str = "") -> List[Dict[str, Any]]:
    """Parse a comp file into a list of comp dicts. Accepts CSV, Excel, or PDF."""
    if hasattr(file_obj_or_bytes, "read"):
        data = file_obj_or_bytes.read()
        if hasattr(file_obj_or_bytes, "name"):
            filename = filename or file_obj_or_bytes.name
    elif isinstance(file_obj_or_bytes, bytes):
        data = file_obj_or_bytes
    else:
        raise ValueError("Pass a file-like object or bytes.")

    lower_name = filename.lower()
    is_pdf = lower_name.endswith(".pdf") or data[:4] == b"%PDF"
    is_excel = (lower_name.endswith(".xlsx") or lower_name.endswith(".xls")
                or lower_name.endswith(".xlsm") or
                (not is_pdf and data[:4] == b"PK\x03\x04"))

    if is_pdf:
        return _parse_pdf_comps(data)
    if is_excel:
        rows = _read_excel_bytes(data)
    else:
        rows = _read_csv_bytes(data)

    header_idx = None
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for v in row if v not in (None, ""))
        if non_empty < 4: continue
        addr_idx = _find_column(row, COLUMN_PATTERNS["address"])
        price_idx = _find_column(row, COLUMN_PATTERNS["sold_price"])
        if addr_idx is not None and price_idx is not None:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find a header row with Address and Price columns.")

    header = rows[header_idx]
    cols = {f: _find_column(header, ps) for f, ps in COLUMN_PATTERNS.items()}

    comps = []
    for row in rows[header_idx + 1:]:
        if all(v in (None, "") for v in row): continue
        def g(field, parser=None):
            i = cols.get(field)
            if i is None or i >= len(row): return None
            v = row[i]
            if parser: return parser(v)
            return v if v not in ("", None) else None

        comp = {
            "address": g("address"), "city": g("city"), "state": g("state"),
            "zip": g("zip"), "beds": g("beds", _parse_int),
            "baths": g("baths"), "sqft": g("sqft", _parse_int),
            "year": g("year", _parse_int), "sold_price": g("sold_price", _parse_money),
            "sold_date": str(g("sold_date") or ""),
            "distance": g("distance"), "notes": g("notes"),
        }
        if not comp["address"] or comp["sold_price"] is None: continue
        comp["dollar_per_sqft"] = (comp["sold_price"] / comp["sqft"]) if comp["sqft"] else None
        comps.append(comp)
    return comps


def suggested_arv(comps: List[Dict[str, Any]], subject_sqft: Optional[int] = None,
                  use_adjusted: bool = False) -> Dict[str, float]:
    """Compute suggested ARV from comp records.

    If ``use_adjusted=True`` and the comp has an ``adjusted_price`` field, that
    is used in place of ``sold_price`` for all calculations.
    """
    price_key = "adjusted_price" if use_adjusted else "sold_price"
    selected = [c for c in comps if c.get(price_key) or c.get("sold_price")]
    if not selected:
        return {"avg_sale": 0, "median_sale": 0, "avg_psf_times_sqft": 0,
                "median_psf_times_sqft": 0, "suggested": 0}
    prices = [c.get(price_key) or c["sold_price"] for c in selected]
    # Compute dollar_per_sqft on the fly if not present (data editor drops it)
    psf = []
    for c in selected:
        price = c.get(price_key) or c.get("sold_price") or 0
        psf_val = c.get("dollar_per_sqft")
        if not psf_val and c.get("sqft") and price:
            psf_val = price / c["sqft"]
        if psf_val:
            psf.append(psf_val)
    avg_sale = sum(prices) / len(prices)
    sorted_p = sorted(prices)
    median_sale = (sorted_p[len(sorted_p)//2] if len(sorted_p) % 2
                   else (sorted_p[len(sorted_p)//2 - 1] + sorted_p[len(sorted_p)//2]) / 2)
    if psf and subject_sqft:
        avg_psf = sum(psf) / len(psf)
        sorted_psf = sorted(psf)
        median_psf = (sorted_psf[len(sorted_psf)//2] if len(sorted_psf) % 2
                      else (sorted_psf[len(sorted_psf)//2 - 1] + sorted_psf[len(sorted_psf)//2]) / 2)
        avg_psf_times_sqft = avg_psf * subject_sqft
        median_psf_times_sqft = median_psf * subject_sqft
    else:
        avg_psf_times_sqft = 0
        median_psf_times_sqft = 0
    methods = [v for v in (avg_sale, median_sale, avg_psf_times_sqft, median_psf_times_sqft) if v > 0]
    suggested = sum(methods) / len(methods) if methods else 0
    return {"avg_sale": avg_sale, "median_sale": median_sale,
            "avg_psf_times_sqft": avg_psf_times_sqft,
            "median_psf_times_sqft": median_psf_times_sqft,
            "suggested": suggested}


# ---------------------------------------------------------------------------
# Comp filter rules — applied AFTER pulling from RentCast
# ---------------------------------------------------------------------------
def filter_comps(comps: List[Dict[str, Any]], subject: Dict[str, Any],
                 rules: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Apply SFR comping filter rules to a list of comps. Returns the subset
    that passes ALL filters.

    Filters applied (any subject value of 0 / None skips that filter):
      - same property_type (strict)
      - distance ≤ comp_max_radius_miles
      - days_old ≤ comp_max_days_old (when sold date is parseable)
      - sqft within ±comp_sqft_tolerance_pct of subject
      - beds within ±comp_beds_tolerance of subject
      - baths within ±comp_baths_tolerance of subject
      - year built within ±comp_year_tolerance of subject
    """
    import datetime as _dt
    subj_sqft = subject.get("sqft", 0) or 0
    subj_beds = subject.get("beds", 0) or 0
    subj_baths = subject.get("baths", 0) or 0
    subj_year = subject.get("year", 0) or 0
    subj_type = (subject.get("property_type") or "").strip()

    max_radius = rules.get("comp_max_radius_miles", 0.5)
    max_days = rules.get("comp_max_days_old", 180)
    sqft_tol = rules.get("comp_sqft_tolerance_pct", 0.25)
    beds_tol = rules.get("comp_beds_tolerance", 1)
    baths_tol = rules.get("comp_baths_tolerance", 0.5)
    year_tol = rules.get("comp_year_tolerance", 15)

    today = _dt.date.today()
    passed = []
    for c in comps:
        # Distance
        if max_radius and (c.get("distance") or 0) > max_radius:
            continue
        # Date (best-effort parse)
        sold_date = c.get("sold_date") or ""
        if max_days and sold_date:
            try:
                d = _dt.datetime.fromisoformat(sold_date.replace("Z", "")).date()
                if (today - d).days > max_days:
                    continue
            except Exception:
                pass  # unparseable date — don't filter out
        # Sqft
        if subj_sqft > 0 and sqft_tol > 0:
            csqft = c.get("sqft") or 0
            if csqft <= 0 or abs(csqft - subj_sqft) > subj_sqft * sqft_tol:
                continue
        # Beds
        if subj_beds > 0 and beds_tol >= 0:
            cbeds = c.get("beds") or 0
            if abs(cbeds - subj_beds) > beds_tol:
                continue
        # Baths
        if subj_baths > 0 and baths_tol >= 0:
            cbaths = c.get("baths") or 0
            if abs(cbaths - subj_baths) > baths_tol:
                continue
        # Year
        if subj_year > 0 and year_tol > 0:
            cyear = c.get("year") or 0
            if cyear > 0 and abs(cyear - subj_year) > year_tol:
                continue
        # Property type (strict match if subject has it set)
        if subj_type and c.get("property_type"):
            # Normalize: drop spaces, compare lowercase
            cnorm = (c["property_type"] or "").strip().lower()
            snorm = subj_type.lower()
            # Allow flexible match — "Single Family" vs "Single Family Residence"
            if not (cnorm in snorm or snorm in cnorm):
                continue
        passed.append(c)
    return passed


# ---------------------------------------------------------------------------
# Comp price adjustments — appraiser-style equivalency math
# ---------------------------------------------------------------------------
def apply_adjustments(comp: Dict[str, Any], subject: Dict[str, Any],
                      adj_table: Dict[str, float]) -> Dict[str, Any]:
    """Return a copy of the comp with an ``adjusted_price`` field added.

    For each amenity that differs between subject and comp, we adjust the comp's
    sold price toward what it would have sold for IF it had matched the subject.
      - subject has feature, comp doesn't → +adj  (comp's value would be higher)
      - subject lacks feature, comp has it → -adj (comp's value includes the premium)
    """
    c = dict(comp)  # shallow copy
    sold = float(c.get("sold_price") or 0)
    if sold <= 0:
        c["adjusted_price"] = 0
        c["adjustments"] = []
        return c

    breakdown = []

    def _adj(name, amount):
        nonlocal sold
        sold += amount
        breakdown.append((name, amount))

    # Pool (subject's pool field is "Yes" / "No"; comp's is bool)
    subj_pool = (subject.get("pool", "No") == "Yes")
    comp_pool = bool(c.get("pool", False))
    pool_v = float(adj_table.get("adj_pool", 0) or 0)
    if pool_v > 0 and subj_pool != comp_pool:
        _adj("Pool", pool_v if subj_pool else -pool_v)

    # Waterfront — manual entry on subject ("No" / "Canal/Lake" / "Ocean")
    subj_wf = (subject.get("waterfront", "No") or "No")
    comp_wf = (c.get("waterfront", "No") or "No")
    canal_v = float(adj_table.get("adj_waterfront_canal", 0) or 0)
    ocean_v = float(adj_table.get("adj_waterfront_ocean", 0) or 0)
    # Score each side: 0=No, 1=Canal/Lake, 2=Ocean
    wf_score = {"No": 0, "Canal/Lake": 1, "Canal": 1, "Lake": 1, "Ocean": 2}
    s_score = wf_score.get(subj_wf, 0)
    c_score = wf_score.get(comp_wf, 0)
    if s_score != c_score:
        # If subject is ocean and comp is no-water: add ocean premium
        # If subject is canal and comp is ocean: subtract (ocean - canal) value
        s_val = {0: 0, 1: canal_v, 2: ocean_v}.get(s_score, 0)
        c_val = {0: 0, 1: canal_v, 2: ocean_v}.get(c_score, 0)
        diff = s_val - c_val
        if diff != 0:
            label = f"Waterfront ({subj_wf} vs {comp_wf})"
            _adj(label, diff)

    # Garage spaces
    subj_garage = int(subject.get("garage_spaces", 0) or 0)
    comp_garage = int(c.get("garage_spaces", 0) or 0)
    if subj_garage != comp_garage:
        g1 = float(adj_table.get("adj_garage_1car", 0) or 0)
        g2 = float(adj_table.get("adj_garage_2car", 0) or 0)
        # Per-car incremental: 1st car = g1; each additional = (g2-g1)
        s_val = subj_garage * g1 if subj_garage <= 1 else g1 + (subj_garage - 1) * (g2 - g1)
        c_val = comp_garage * g1 if comp_garage <= 1 else g1 + (comp_garage - 1) * (g2 - g1)
        diff = s_val - c_val
        if diff != 0:
            _adj(f"Garage ({subj_garage} vs {comp_garage})", diff)

    # Bedroom delta (informational — sqft tolerance usually handles this, but
    # appraisers still do explicit bedroom adjustments)
    subj_beds = int(subject.get("beds", 0) or 0)
    comp_beds = int(c.get("beds", 0) or 0)
    bed_v = float(adj_table.get("adj_extra_bedroom", 0) or 0)
    if bed_v > 0 and subj_beds != comp_beds:
        _adj(f"Bedrooms ({subj_beds} vs {comp_beds})",
             (subj_beds - comp_beds) * bed_v)

    # Half-bath delta (using ceiling for any non-integer bath count)
    subj_baths = float(subject.get("baths", 0) or 0)
    comp_baths = float(c.get("baths", 0) or 0)
    halfbath_v = float(adj_table.get("adj_extra_half_bath", 0) or 0)
    bath_diff_halves = round((subj_baths - comp_baths) * 2)
    if halfbath_v > 0 and bath_diff_halves != 0:
        _adj(f"Baths ({subj_baths} vs {comp_baths})",
             bath_diff_halves * halfbath_v)

    c["adjusted_price"] = sold
    c["adjustments"] = breakdown
    # Recompute $/sqft based on adjusted price
    if c.get("sqft"):
        c["adjusted_dollar_per_sqft"] = sold / c["sqft"]
    return c


def adjust_all(comps: List[Dict[str, Any]], subject: Dict[str, Any],
               adj_table: Dict[str, float]) -> List[Dict[str, Any]]:
    """Apply adjustments to every comp in a list."""
    return [apply_adjustments(c, subject, adj_table) for c in comps]
