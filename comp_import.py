"""Comp import — parse MLS exports (CSV/Excel/PDF) into structured comp records.

Supports three input formats:
  - CSV (from Matrix, Flexmls, etc.)
  - Excel (.xlsx, .xls, .xlsm)
  - PDF (Comparable Sales Reports from data providers like PropStream)

Returns a list of comp dicts with normalized keys.
"""
import re, csv, io
from typing import Optional, List, Dict, Any


COLUMN_PATTERNS = {
    "address":   ["property address", "full address", "site address", "street address", "address"],
    "city":      ["city", "municipality"],
    "state":     ["state"],
    "zip":       ["zip code", "zipcode", "postal code", "postal", "zip"],
    "beds":      ["bedrooms", "bed rooms", "beds", "br", "# beds"],
    "baths":     ["total baths", "bathrooms", "baths", "ba", "# baths", "full baths"],
    "sqft":      ["total living area", "living area sqft", "living sqft", "living area",
                  "sq ft", "sqft", "square feet", "approx sqft"],
    "year":      ["year built", "yr built", "year"],
    "sold_price":["close price", "closed price", "sold price", "sale price", "sold for", "price"],
    "sold_date": ["close date", "closed date", "sold date", "sale date", "closing date"],
    "distance":  ["approximate distance", "distance", "miles", "mi"],
    "notes":     ["public remarks", "remarks", "notes", "comments"],
}


def _normalize(s):
    if s is None: return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _find_column(header, patterns):
    norm = [_normalize(h) for h in header]
    for p in patterns:
        pn = _normalize(p)
        for i, h in enumerate(norm):
            if h == pn: return i
    for p in patterns:
        pn = _normalize(p)
        for i, h in enumerate(norm):
            if pn and (pn in h or (len(h) >= 3 and h in pn)):
                return i
    return None


def _parse_money(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try: return float(s) if s else None
    except ValueError: return None


def _parse_int(v):
    n = _parse_money(v)
    return int(n) if n is not None else None


def _parse_first_int(v):
    """Extract the FIRST integer from a string. Handles 'X,XXX Δ -YYY' format."""
    if v is None or v == "": return None
    if isinstance(v, int): return v
    if isinstance(v, float): return int(v)
    m = re.search(r"[\d,]+", str(v))
    if not m: return None
    try: return int(m.group(0).replace(",", ""))
    except ValueError: return None


def _read_csv_bytes(data: bytes) -> List[List]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_excel_bytes(data: bytes) -> List[List]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


# ============================================================================
# PDF PARSER — coordinate-based, handles column-tabular CMA reports
# ============================================================================
def _cluster_x_positions(xs, tolerance=25):
    """Cluster X coordinates into column centers."""
    if not xs:
        return []
    sorted_xs = sorted(set(xs))
    clusters = [[sorted_xs[0]]]
    for x in sorted_xs[1:]:
        if x - clusters[-1][-1] <= tolerance:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c) / len(c) for c in clusters]


def _assign_to_column(x, column_anchors, max_extent=120):
    """Return index of column whose anchor is just to the left of x.
    Treats column_anchors as left-edges of columns. A word at x belongs
    to column i where anchor[i] <= x < anchor[i+1] (or unbounded for last col).
    """
    if not column_anchors:
        return None
    # If word is before the first column start, no column
    if x < column_anchors[0] - 10:
        return None
    for i in range(len(column_anchors) - 1):
        if column_anchors[i] - 10 <= x < column_anchors[i+1] - 10:
            return i
    # Last column — accept words up to max_extent beyond the anchor
    last = column_anchors[-1]
    if x >= last - 10 and x < last + max_extent:
        return len(column_anchors) - 1
    return None


def _parse_pdf_comps(data: bytes) -> List[Dict[str, Any]]:
    """Parse a Comparable Sales Report style PDF (from PropStream and similar)."""
    import pdfplumber

    all_comps = []  # accumulated across pages
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        # Concatenate words from all pages that contain a "Comparable List"
        for page in pdf.pages:
            text = page.extract_text() or ""
            if "Comparable List" not in text and "Comparable Sales Report" not in text:
                # also skip pages that don't have property tables
                continue
            words = page.extract_words(use_text_flow=True)
            if not words:
                continue

            # Group words by row (Y coordinate buckets)
            from collections import defaultdict
            rows = defaultdict(list)
            for w in words:
                key = round(w["top"] / 5) * 5
                rows[key].append(w)

            # Discover column anchors from rows that have ONE token per column
            # (Bed, Bath, Stories, AVM all have single-word numeric values).
            # Rows like "Lot Size" (multi-word values) or "Type" (multi-word
            # values) would produce spurious extra column anchors.
            candidate_xs = []
            for y in sorted(rows):
                line_words = sorted(rows[y], key=lambda w: w["x0"])
                if not line_words: continue
                first_text = line_words[0]["text"]
                if first_text in ("Bed", "Bath", "Stories", "AVM"):
                    for w in line_words[1:]:
                        candidate_xs.append(w["x0"])

            if not candidate_xs:
                continue

            column_centers = _cluster_x_positions(candidate_xs, tolerance=60)
            # First column is subject, the rest are comps
            if len(column_centers) < 2:
                continue

            # Now extract field values per column
            # Build map of field_name → {col_idx: [value tokens]}
            fields = defaultdict(lambda: defaultdict(list))

            current_field = None
            for y in sorted(rows):
                line_words = sorted(rows[y], key=lambda w: w["x0"])
                if not line_words: continue
                first_word = line_words[0]
                first_text = first_word["text"]
                first_x = first_word["x0"]

                # Field labels are in the leftmost area (x < ~120)
                if first_x < column_centers[0] - 20:
                    # Concatenate multi-word labels (e.g., "Last Sold", "Year Built")
                    label_words = []
                    value_words = []
                    for w in line_words:
                        if w["x0"] < column_centers[0] - 20:
                            label_words.append(w["text"])
                        else:
                            value_words.append(w)
                    label = " ".join(label_words).strip()
                    current_field = label
                    # Distribute value_words to columns
                    for w in value_words:
                        col = _assign_to_column(w["x0"], column_centers)
                        if col is not None:
                            fields[current_field][col].append(w["text"])
                else:
                    # No label — continuation of previous field (e.g., date under price)
                    if current_field:
                        for w in line_words:
                            col = _assign_to_column(w["x0"], column_centers)
                            if col is not None:
                                fields[current_field][col].append(w["text"])

            # Now reconstruct comp records (skip subject = column 0)
            num_comps = len(column_centers) - 1
            for comp_idx in range(num_comps):
                col = comp_idx + 1
                def get(field_name):
                    tokens = fields.get(field_name, {}).get(col, [])
                    return " ".join(tokens).strip() if tokens else None

                # Address is in the first 3 lines under "Subject Property" / each property header
                # In this layout, the first field with a value is something like "879 Ne 158Th St"
                # but the field label section doesn't have a name for it — the lines before "Bed".
                # Instead, look for any field whose key looks address-y.
                # In the PropStream format, the address is stored under "Subject Property" (col 0)
                # and the comp addresses are positional. Let's reconstruct from the header rows.

                # Sold price + date
                last_sold = get("Last Sold")
                price = None
                date = None
                if last_sold:
                    m = re.search(r"\$[\d,]+", last_sold)
                    if m: price = _parse_money(m.group())
                    m = re.search(r"\d{1,2}/\d{1,2}/\d{4}", last_sold)
                    if m: date = m.group()

                sqft = _parse_first_int(get("Square Feet"))
                beds = _parse_first_int(get("Bed"))
                baths = get("Bath")  # keep as text "2.0"
                year = _parse_int(get("Year Built"))
                listing_status = get("Listing Status")

                # Skip if no price (e.g., a non-sold listing)
                if price is None:
                    continue

                # If listing_status indicates a sale, prefer that, otherwise still take it
                comp = {
                    "address": None,  # filled below from address rows
                    "city": None, "state": None, "zip": None,
                    "beds": beds, "baths": baths, "sqft": sqft, "year": year,
                    "sold_price": price, "sold_date": date,
                    "distance": None, "notes": listing_status or "",
                    "dollar_per_sqft": (price / sqft) if (price and sqft) else None,
                }
                all_comps.append((col, comp))

            # Extract addresses by reading rows ABOVE "Bed" — address spans 3 lines
            # (street, city/state, zip). Need a wider Y window to capture all 3.
            sorted_ys = sorted(rows)
            bed_y = None
            for y in sorted_ys:
                texts = [w["text"] for w in rows[y]]
                if "Bed" in texts:
                    bed_y = y; break

            if bed_y is not None:
                # Find the Y range of address rows — between the "Subject Property" label
                # and the "Bed" row. Typically 3 lines of address data per property.
                addr_rows = []
                for y in sorted_ys:
                    if y >= bed_y: continue
                    texts = [w["text"] for w in rows[y]]
                    # Skip "Subject Property" label line and "Comparable List" header
                    if any(t in ("Subject", "Property", "Comparable", "List") for t in texts):
                        continue
                    addr_rows.append(y)
                # Take the last 4 rows before Bed (covers 3-line addresses with safety margin)
                addr_rows = addr_rows[-4:] if len(addr_rows) > 4 else addr_rows

                # For each comp column, collect address tokens by row order
                addr_by_col = defaultdict(list)
                for y in addr_rows:
                    for w in sorted(rows[y], key=lambda w: w["x0"]):
                        col = _assign_to_column(w["x0"], column_centers)
                        if col is not None and col >= 1:
                            addr_by_col[col].append(w["text"])

                for col, addr_tokens in addr_by_col.items():
                    full = " ".join(addr_tokens)
                    # Zip = the LAST 5-digit number in the address (street numbers
                    # like '16200' can match earlier; we want the trailing zip).
                    zip_matches = re.findall(r"\b(\d{5})\b", full)
                    zip_code = zip_matches[-1] if zip_matches else None
                    # State = 2-letter state abbreviation immediately before the zip
                    # (handles addresses like '16200 NE 8 CT ... FL 33162' where NE
                    # is the directional, not the state).
                    state_code = None
                    state_pattern = (r"\b(FL|GA|NC|SC|TX|CA|NY|NJ|PA|VA|MD|MA|TN|OH|"
                                     r"MI|IL|IN|KY|AL|AR|LA|MS|MO|OK|KS|CO|UT|AZ|NV|"
                                     r"OR|WA|WI|MN|IA|NE|SD|ND|MT|WY|ID|NM|HI|AK|ME|"
                                     r"VT|NH|RI|CT|DE|WV)\b\s+\d{5}\b")
                    sm = re.search(state_pattern, full, re.I)
                    if sm: state_code = sm.group(1).upper()
                    for c2, comp in all_comps:
                        if c2 == col and comp.get("address") is None:
                            comp["address"] = full
                            comp["zip"] = zip_code
                            comp["state"] = state_code
                            break

    return [c for _, c in all_comps]


def parse_comp_file(file_obj_or_bytes, filename: str = "") -> List[Dict[str, Any]]:
    """Parse a comp file into a list of comp dicts. Accepts CSV, Excel, or PDF."""
    if hasattr(file_obj_or_bytes, "read"):
        data = file_obj_or_bytes.read()
        if hasattr(file_obj_or_bytes, "name"):
            filename = filename or file_obj_or_bytes.name
    elif isinstance(file_obj_or_bytes, bytes):
        data = file_obj_or_bytes
    else:
        raise ValueError("Pass a file-like object or bytes.")

    lower_name = filename.lower()
    is_pdf = lower_name.endswith(".pdf") or data[:4] == b"%PDF"
    is_excel = (lower_name.endswith(".xlsx") or lower_name.endswith(".xls")
                or lower_name.endswith(".xlsm") or
                (not is_pdf and data[:4] == b"PK\x03\x04"))

    if is_pdf:
        return _parse_pdf_comps(data)
    if is_excel:
        rows = _read_excel_bytes(data)
    else:
        rows = _read_csv_bytes(data)

    # CSV/Excel branch — find header row
    header_idx = None
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for v in row if v not in (None, ""))
        if non_empty < 4: continue
        addr_idx = _find_column(row, COLUMN_PATTERNS["address"])
        price_idx = _find_column(row, COLUMN_PATTERNS["sold_price"])
        if addr_idx is not None and price_idx is not None:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find a header row with Address and Price columns.")

    header = rows[header_idx]
    cols = {f: _find_column(header, ps) for f, ps in COLUMN_PATTERNS.items()}

    comps = []
    for row in rows[header_idx + 1:]:
        if all(v in (None, "") for v in row): continue
        def g(field, parser=None):
            i = cols.get(field)
            if i is None or i >= len(row): return None
            v = row[i]
            if parser: return parser(v)
            return v if v not in ("", None) else None

        comp = {
            "address": g("address"), "city": g("city"), "state": g("state"),
            "zip": g("zip"), "beds": g("beds", _parse_int),
            "baths": g("baths"), "sqft": g("sqft", _parse_int),
            "year": g("year", _parse_int), "sold_price": g("sold_price", _parse_money),
            "sold_date": str(g("sold_date") or ""),
            "distance": g("distance"), "notes": g("notes"),
        }
        if not comp["address"] or comp["sold_price"] is None: continue
        comp["dollar_per_sqft"] = (comp["sold_price"] / comp["sqft"]) if comp["sqft"] else None
        comps.append(comp)
    return comps


def suggested_arv(comps: List[Dict[str, Any]], subject_sqft: Optional[int] = None) -> Dict[str, float]:
    selected = [c for c in comps if c.get("sold_price")]
    if not selected:
        return {"avg_sale": 0, "median_sale": 0, "avg_psf_times_sqft": 0,
                "median_psf_times_sqft": 0, "suggested": 0}
    prices = [c["sold_price"] for c in selected]
    psf = [c["dollar_per_sqft"] for c in selected if c.get("dollar_per_sqft")]
    avg_sale = sum(prices) / len(prices)
    sorted_p = sorted(prices)
    median_sale = (sorted_p[len(sorted_p)//2] if len(sorted_p) % 2
                   else (sorted_p[len(sorted_p)//2 - 1] + sorted_p[len(sorted_p)//2]) / 2)
    if psf and subject_sqft:
        avg_psf = sum(psf) / len(psf)
        sorted_psf = sorted(psf)
        median_psf = (sorted_psf[len(sorted_psf)//2] if len(sorted_psf) % 2
                      else (sorted_psf[len(sorted_psf)//2 - 1] + sorted_psf[len(sorted_psf)//2]) / 2)
        avg_psf_times_sqft = avg_psf * subject_sqft
        median_psf_times_sqft = median_psf * subject_sqft
    else:
        avg_psf_times_sqft = 0
        median_psf_times_sqft = 0
    methods = [v for v in (avg_sale, median_sale, avg_psf_times_sqft, median_psf_times_sqft) if v > 0]
    suggested = sum(methods) / len(methods) if methods else 0
    return {"avg_sale": avg_sale, "median_sale": median_sale,
            "avg_psf_times_sqft": avg_psf_times_sqft,
            "median_psf_times_sqft": median_psf_times_sqft,
            "suggested": suggested}
